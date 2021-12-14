from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import selenium.common.exceptions
import time
import openpyxl
from urllib.parse import quote_plus
from ajeitaTelefones import ajeitaTels
from webdriver_manager.chrome import ChromeDriverManager
options = webdriver.ChromeOptions()
# path = "user-data-dir=data"
# options.add_argument(path)

ajeitaTels()

driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
driver.get('https://web.whatsapp.com/')
time.sleep(10)

# verificador_entrada = WebDriverWait(driver, 45).until(ec.presence_of_element_located((By.XPATH, "//div[@class='landing-main']")))
wb = openpyxl.load_workbook('basededados.xlsx')
sheet = wb.active
file = open('msg.txt', 'r', encoding='utf-8')
file1 = open('link.txt', 'r', encoding='utf-8')
link = file1.read()
mensagem2 = file.read()
# mexer no range e loads
for i in range(1, sheet.max_row):
    if sheet.cell(row=i+1, column=2).value is None and sheet.cell(row=i+2, column=2).value is None:
        break
    elif sheet.cell(row=i+1, column=5).value == 'v':
        print("Já enviamos para esse numero")
        continue
    else:
        nome = sheet.cell(row=i + 1, column=1).value.split()
        mensagem3 = mensagem2.replace("NOME", nome[0])
        encoded_msg = quote_plus(mensagem3, safe='?!,@')
        telefone = str(sheet.cell(row=i + 1, column=2).value)
        url = 'https://web.whatsapp.com/send?phone=55' + telefone + '&text=' +encoded_msg

        driver.get(url)
        print("Entrando na conversa...")
        driver.execute_script("window.onbeforeunload = function() {};")
        try:
            #time.sleep(7)
            #inp_xpath = '//div[@class="_13NKt copyable-text selectable-text"][@contenteditable="true"][@data-tab="9"]'
            #input_box = driver.find_element_by_xpath(inp_xpath)
            #time.sleep(2)
            #input_box.send_keys(Keys.ENTER)
            #time.sleep(2)   
            #input_box.send_keys(link+Keys.ENTER)

            elemento = WebDriverWait(driver, 13).until(ec.presence_of_element_located((By.XPATH, "//div[@class='_13NKt copyable-text selectable-text'][@data-tab='9']")))
            time.sleep(2)
            elemento.send_keys(Keys.RETURN)
            elemento.send_keys(link)
            sheet.cell(row=i + 1, column=5).value = 'v'
            elemento.send_keys(Keys.RETURN)
            time.sleep(3)
            print("Enviando...")
            wb.save('basededados.xlsx')
            continue
        except selenium.common.exceptions.TimeoutException:
            try:
                print("Não encontrou o elemento na 1a tentativa, tentando novamente...")
                elemento = WebDriverWait(driver, 13).until(ec.presence_of_element_located((By.XPATH, "//div[@class='_13NKt copyable-text selectable-text'][@data-tab='3']")))
                time.sleep(2)
                msg1_box = driver.switch_to.active_element
                msg1_box.send_keys(Keys.RETURN)
                msg1_box.send_keys(link)
                sheet.cell(row=i + 1, column=5).value = 'v'
                msg1_box.send_keys(Keys.RETURN)
                time.sleep(3)
                print(2)
                wb.save('basededados.xlsx')
                continue
            except selenium.common.exceptions.TimeoutException:
                try:
                    print("Não encontrou o elemento na 2a tentativa, tentando novamente...")
                    elemento = WebDriverWait(driver, 13).until(ec.presence_of_element_located((By.XPATH, "//div[@class='_13NKt copyable-text selectable-text'][@data-tab='3']")))
                    time.sleep(2)
                    msg1_box = driver.switch_to.active_element
                    msg1_box.send_keys(Keys.RETURN)
                    msg1_box.send_keys(link)
                    sheet.cell(row=i + 1, column=5).value = 'v'
                    msg1_box.send_keys(Keys.RETURN)
                    time.sleep(5)
                    print("Enviando...")
                    wb.save('basededados.xlsx')
                    continue
                except selenium.common.exceptions.TimeoutException:
                    sheet.cell(row=i + 1, column=4).value = 'v'
                    print("Conexão provavelmente lenta ou desconectada, mandar msg pro Matheus Domingos")
                    continue
        except selenium.common.exceptions.InvalidElementStateException:
            try:
                print(7)
                elemento = WebDriverWait(driver, 13).until(ec.presence_of_element_located((By.XPATH, "//div[@class='_13NKt copyable-text selectable-text'][@data-tab='3']")))
                time.sleep(2)
                msg1_box = driver.switch_to.active_element
                msg1_box.send_keys(Keys.RETURN)
                msg1_box.send_keys(link)
                sheet.cell(row=i + 1, column=5).value = 'v'
                msg1_box.send_keys(Keys.RETURN)
                time.sleep(5)
                print("Enviando...")
                wb.save('basededados.xlsx')
                continue
            except selenium.common.exceptions.TimeoutException:
                sheet.cell(row=i + 1, column=4).value = 'v'
                print("Conexão provavelmente lenta ou desconectada, mandar msg para o Matheus Domingos")
                continue
print("wppapi finalizado!")