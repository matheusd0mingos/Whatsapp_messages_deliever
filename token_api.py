from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import selenium.common.exceptions
import time
import openpyxl
# Time outs muito aumentados por possível problema de conexão do andré
# Ajeitei o problema apresentado pela Paula quando era preciso começar por alguém no meio da lista, pequena mudança em como ele processa as informações
# fiz um pequeno estudo sobre envio de documentos pelo selenium,
driver = webdriver.Chrome(executable_path='chromedriver.exe')
driver.get('https://web.whatsapp.com/')
time.sleep(15)
wb = openpyxl.load_workbook('basededados.xlsx')
sheet = wb.active
file = open('msg.txt', 'r', encoding='utf-8')
file2 = open('msg2.txt', 'r', encoding='utf-8')
mensagem2 = file.read()
mensagem22 = file2.read()
# mexer no range e loads
for i in range(1, sheet.max_row):
    print(sheet.cell(row=i+1, column=1).value)
    if sheet.cell(row=i+1, column=2).value is None and sheet.cell(row=i+2, column=2).value is None:
        break
    elif sheet.cell(row=i+1, column=5).value == 'x':
        print(13)
        continue
    elif sheet.cell(row=i+1, column=5).value == 'v':
        print(14)
        continue
    else:
        nome_completo = sheet.cell(row=i+1, column=1).value
        print(nome_completo)
        nome = nome_completo.split()[0]
        token = sheet.cell(row=i+1, column=4).value
        print(token)
        mensagem3 = mensagem2.replace("NOME", nome).replace("CARGO", token)
        mensagem33 = mensagem22.replace("NOME", nome).replace("CARGO", token)
        driver.get('https://web.whatsapp.com/send?phone=55' + str(sheet.cell(row=i + 1, column=2).value))
        print(11)
        driver.execute_script("window.onbeforeunload = function() {};")
        msg_list = mensagem3.split('\n')
        msg2_list = mensagem33.split('\n')
        try:
            print(4)
            elemento = WebDriverWait(driver, 15).until(ec.presence_of_element_located((By.XPATH, "//div[@class='_3FRCZ copyable-text selectable-text'][@data-tab='1']")))
            msg1_box = driver.switch_to.active_element
            for msg in msg_list:
                msg1_box.send_keys(msg)
                msg1_box.send_keys(Keys.SHIFT + Keys.RETURN)
            sheet.cell(row=i + 1, column=5).value = 'v'
            msg1_box.send_keys(Keys.RETURN)
            for msg in msg2_list:
                msg1_box.send_keys(msg)
                msg1_box.send_keys(Keys.SHIFT + Keys.RETURN)
            sheet.cell(row=i + 1, column=5).value = 'v'
            msg1_box.send_keys(Keys.RETURN)
            time.sleep(2)
            print(2)
            wb.save('basededados.xlsx')
            continue
        except selenium.common.exceptions.TimeoutException:
            try:
                print(5)
                elemento = WebDriverWait(driver, 15).until(ec.presence_of_element_located((By.XPATH, "//div[@class='_3FRCZ copyable-text selectable-text'][@data-tab='1']")))
                msg1_box = driver.switch_to.active_element
                for msg in msg_list:
                    msg1_box.send_keys(msg)
                    msg1_box.send_keys(Keys.SHIFT + Keys.RETURN)
                sheet.cell(row=i + 1, column=5).value = 'v'
                msg1_box.send_keys(Keys.RETURN)
                for msg in msg2_list:
                    msg1_box.send_keys(msg)
                    msg1_box.send_keys(Keys.SHIFT + Keys.RETURN)
                sheet.cell(row=i + 1, column=5).value = 'v'
                msg1_box.send_keys(Keys.RETURN)
                time.sleep(4)
                print(2)
                wb.save('basededados.xlsx')
                continue
            except selenium.common.exceptions.TimeoutException:
                sheet.cell(row=i+1, column=5).value = 'x'
                wb.save('basededados.xlsx')
print("wppapi finalizado!")
