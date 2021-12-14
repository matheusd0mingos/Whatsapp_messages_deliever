from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import selenium.common.exceptions
import time
import openpyxl

driver = webdriver.Chrome(executable_path='chromedriver.exe')
driver.get('https://web.whatsapp.com/')
time.sleep(20)
wb = openpyxl.load_workbook('basededados.xlsx')
sheet = wb.active
for i in range(1, sheet.max_row):
    if sheet.cell(row=i+1, column=2).value is None and sheet.cell(row=i+2, column=2).value is None:
        break
    if sheet.cell(row=i+1, column=4).value == 'v' or sheet.cell(row=i+1, column=4).value == 'x':
        continue
    else:
        driver.get('https://web.whatsapp.com/send?phone=55' + str(sheet.cell(row=i+1, column=2).value))
        try:
            popup_invalid_tel = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@class='_3lLzD']")))
            invalido = driver.find_element_by_xpath("//div[@class='_3lLzD']")
            if "invŕ¸Łŕ¸lido" in invalido.text:
                sheet.cell(row=i+1, column=4).value = 'x'
                print(1)
                wb.save('basededados.xlsx')
                continue
            else:
                time.sleep(0.5)
                msg1_box = driver.switch_to.active_element
                sheet.cell(row=i+1, column=4).value = 'v'
                print(2)
                wb.save('basededados.xlsx')
                continue
        except selenium.common.exceptions.TimeoutException:
            time.sleep(0.75)
            msg1_box = driver.switch_to.active_element
            sheet.cell(row=i+1, column=4).value = 'v'
            print(3)
            wb.save('basededados.xlsx')
            continue
        except selenium.common.exceptions.StaleElementReferenceException:
            time.sleep(0.75)
            msg1_box = driver.switch_to.active_element
            sheet.cell(row=i + 1, column=4).value = 'v'
            print(3)
            wb.save('basededados.xlsx')
            continue
        except selenium.common.exceptions.NoSuchElementException:
            time.sleep(0.75)
            msg1_box = driver.switch_to.active_element
            sheet.cell(row=i + 1, column=4).value = 'v'
            print(3)
            wb.save('basededados.xlsx')
            continue
print("verifierwpp finalizado!")



